from openpyxl import Workbook, load_workbook

wb = load_workbook('./excel/transactions.xlsx')

ws = wb.active
ws['D{}'.format(1)] = 'Corrected'
for i in range(2, ws.max_row + 1):
    new_nb = ws[f'C{i}'].value.split(' ')
    new_nb = float(new_nb[0].replace(',','.'))
    calculated_value = new_nb * 0.9
    calculated_value = round(calculated_value, 2)
    calculated_value = str(calculated_value) + ' €'
    calculated_value = calculated_value.replace('.', ',')
    ws['D{}'.format(i)] = calculated_value
    

wb.save('./excel/transactions_modified.xlsx')